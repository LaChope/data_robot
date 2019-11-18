import os, sys, subprocess, getpass, logging, configparser, fnmatch
import sqlite3, openpyxl, json, requests
import pandas as pd
from collections import OrderedDict
from datetime import datetime
from datetime import timedelta
from os.path import join as pjoin
from openpyxl.utils import get_column_letter


# --- Global Variables ----------------------------------------------------------------------------
INIFILENAME = "data_ETL_service.ini"

# --- Classes -------------------------------------------------------------------------------------
class Singleton(type):
    _instances = {}
    def __call__(cls, *args, **kwargs):
        if cls not in cls._instances:
            cls._instances[cls] = super(Singleton, cls).__call__(*args, **kwargs)
        return cls._instances[cls]

class Logger(object, metaclass=Singleton):
    logger = logging.getLogger('logger')
    logger.setLevel(logging.INFO)

    def get_logger(self, initStruct):
        logFilePath = pjoin(initStruct['working_folder'], initStruct['logFilename'])
        formatter = logging.Formatter('%(asctime)s - %(message)s')

        if initStruct['logToConsole'] == 'True':
            log2console = logging.StreamHandler()
            log2console.setLevel(logging.INFO)
            log2console.setFormatter(formatter)
            self.logger.addHandler(log2console)
        
        if initStruct['logToFile'] == 'True':
            log2file = logging.FileHandler(logFilePath)
            log2file.setLevel(logging.INFO)
            log2file.setFormatter(formatter)
            self.logger.addHandler(log2file)

        return self.logger

class DatabaseOperations():

    def __new__(cls):
        self = object.__new__(cls)
        return self

    def openDatabase(self, dbFilePath):
        sqliteDb = sqlite3.connect(dbFilePath)
        sqliteDb.text_factory = str
        return sqliteDb

    def getTableColumns(self, db, table):
        query = 'SELECT * FROM ' + table
        selectData = db.execute(query)
        issueColnames = selectData.description
        return issueColnames

    def getMainTasksCreatedInGivenCW(self, cursor, cwStart, cwEnd, pd):
        query = "SELECT Project.Name, Issue.IssueKey, Issue.IssueType, Issue.Assignee, Issue.Status, Issue.StoryPoints, Issue.Created, Issue.Resolved \
            FROM issue JOIN Project ON Issue.ProjectID = Project.ProjectID \
            WHERE Created BETWEEN " + "'" + cwStart + "'" + " AND " + "'" + cwEnd + "'"
        cursor.execute(query)
        sqlData = cursor.fetchall()
        dataset = pd.DataFrame(data=sqlData)
        dataset = dataset.fillna('')
        return dataset

    def getMainTasksClosedInGivenCW(self, cursor, cwStart, cwEnd, pd):
        query = "SELECT Project.Name, Issue.IssueKey, Issue.IssueType, Issue.Assignee, Issue.Status, Issue.StoryPoints, Issue.Created, Issue.Resolved \
                FROM issue JOIN Project ON Issue.ProjectID = Project.ProjectID \
                WHERE Resolved BETWEEN " + "'" + cwStart + "'" + " AND " + "'" + cwEnd + "'"
        cursor.execute(query)
        sqlData = cursor.fetchall()
        dataset = pd.DataFrame(data=sqlData)
        dataset = dataset.fillna('')
        return dataset

    def getMainTasksInProgress(self, cursor, pd):
        query = "SELECT Project.Name, Issue.IssueKey, Issue.IssueType, Issue.Assignee, Issue.Status, Issue.StoryPoints, Issue.Created, Issue.Resolved \
                FROM issue JOIN Project ON Issue.ProjectID = Project.ProjectID \
                WHERE Issue.Status = 'In Progress'"
        cursor.execute(query)
        sqlData = cursor.fetchall()
        dataset = pd.DataFrame(data=sqlData)
        dataset = dataset.fillna('')
        return dataset

    def getSubtasksStoryPointsCreatedInGivenCW(self, cursor, cwStart, cwEnd, pd):
        query = "SELECT Project.Name, Issue.IssueType, Issue.issueKey, Issue.Status, Issue.StoryPoints, SUM (SubTask.StoryPoints) \
                FROM Issue \
                JOIN SubTask ON Issue.IssueID = SubTask.IssueID \
                JOIN Project ON Issue.ProjectID = Project.ProjectID \
                WHERE Issue.Created BETWEEN " + "'" + cwStart + "'" + " AND " + "'" + cwEnd + "' OR SubTask.Created BETWEEN " + "'" + cwStart + "'" + " AND " + "'" + cwEnd + "' \
                GROUP BY Issue.issueKey"
        cursor.execute(query)
        sqlData = cursor.fetchall()
        dataset = pd.DataFrame(data=sqlData)
        dataset = dataset.fillna('')
        return dataset

    # overall ----------------------------------------------------------------------------
    def getSubtasksStoryPointsClosedInGivenCW(self, cursor, cwStart, cwEnd, pd):
        query = "SELECT Project.Name, SUM (SubTask.StoryPoints) \
                FROM Issue \
                JOIN Project ON Issue.ProjectID = Project.ProjectID \
                JOIN SubTask ON Issue.IssueID = SubTask.IssueID \
                WHERE SubTask.Resolved BETWEEN " + "'" + cwStart + "'" + " AND " + "'" + cwEnd + "' \
                    AND Issue.SubTasks <> 'None' \
                    AND (Issue.Assignee NOT LIKE '%Jan%' OR Issue.Assignee NOT LIKE '%Andrei%')"
        cursor.execute(query)
        sqlData = cursor.fetchall()
        dataset = pd.DataFrame(data=sqlData)
        dataset = dataset.fillna('')
        return dataset

    def getMainTasksStoryPointsClosedInGivenCW(self, cursor, cwStart, cwEnd, pd):
        query = "SELECT Project.Name, SUM (SubTask.StoryPoints) \
                FROM Issue \
                JOIN Project ON Issue.ProjectID = Project.ProjectID \
                JOIN SubTask ON Issue.IssueID = SubTask.IssueID \
                WHERE Issue.Resolved BETWEEN " + "'" + cwStart + "'" + " AND " + "'" + cwEnd + "' \
                    AND Issue.SubTasks = 'None' \
                    AND (Issue.Assignee NOT LIKE '%Jan%' OR Issue.Assignee NOT LIKE '%Andrei%')"
        cursor.execute(query)
        sqlData = cursor.fetchall()
        dataset = pd.DataFrame(data=sqlData)
        dataset = dataset.fillna('')
        return dataset

    def getSubtasksStoryPointsOpen(self, cursor, pd):
        query = "SELECT Project.Name, SUM (SubTask.StoryPoints) \
                FROM Issue \
                JOIN Project ON Issue.ProjectID = Project.ProjectID \
                JOIN SubTask ON Issue.IssueID = SubTask.IssueID \
                WHERE Issue.Status <> 'Closed' \
                    AND SubTask.Status <> 'Closed' \
                    AND Issue.SubTasks <> 'None' \
                    AND (Issue.Assignee NOT LIKE '%Jan%' OR Issue.Assignee NOT LIKE '%Andrei%')"
        cursor.execute(query)
        sqlData = cursor.fetchall()
        dataset = pd.DataFrame(data=sqlData)
        dataset = dataset.fillna('')
        return dataset
    
    def getMainTasksStoryPointsOpen(self, cursor, pd):
        query = "SELECT Project.Name, SUM (SubTask.StoryPoints) \
                FROM Issue \
                JOIN Project ON Issue.ProjectID = Project.ProjectID \
                JOIN SubTask ON Issue.IssueID = SubTask.IssueID \
                WHERE Issue.Status <> 'Closed' \
                    AND SubTask.Status <> 'Closed' \
                    AND Issue.SubTasks = 'None' \
                    AND (Issue.Assignee NOT LIKE '%Jan%' OR Issue.Assignee NOT LIKE '%Andrei%')"
        cursor.execute(query)
        sqlData = cursor.fetchall()
        dataset = pd.DataFrame(data=sqlData)
        dataset = dataset.fillna('')
        return dataset

    def getMainTasksStoryPointsBlocked(self, cursor, pd):
        query = "SELECT Project.Name, SUM(Issue.StoryPoints) \
                FROM Issue \
                JOIN Project ON Issue.ProjectID = Project.ProjectID \
                WHERE Issue.Flagged LIKE '%Impediment%' \
                    AND Issue.Status != 'Closed' \
                    AND (Issue.Assignee NOT LIKE '%Jan%' OR Issue.Assignee NOT LIKE '%Andrei%')"
        cursor.execute(query)
        sqlData = cursor.fetchall()
        dataset = pd.DataFrame(data=sqlData)
        dataset = dataset.fillna('')
        return dataset
    
    
    # IDC5 ----------------------------------------------------------------------------
    def getSubtasksStoryPointsClosedInGivenCW_IDC5(self, cursor, cwStart, cwEnd, pd):
        query = "SELECT Project.Name, SUM (SubTask.StoryPoints) \
                FROM Issue \
                JOIN Project ON Issue.ProjectID = Project.ProjectID \
                JOIN SubTask ON Issue.IssueID = SubTask.IssueID \
                WHERE Project.Name = 'DAI DASy IDC5' \
                    AND SubTask.Resolved BETWEEN " + "'" + cwStart + "'" + " AND " + "'" + cwEnd + "' \
                    AND Issue.SubTasks <> 'None' \
                    AND (Issue.Assignee NOT LIKE '%Jan%' OR Issue.Assignee NOT LIKE '%Andrei%')"
        cursor.execute(query)
        sqlData = cursor.fetchall()
        dataset = pd.DataFrame(data=sqlData)
        dataset = dataset.fillna('')
        return dataset

    def getMainTasksStoryPointsClosedInGivenCW_IDC5(self, cursor, cwStart, cwEnd, pd):
        query = "SELECT Project.Name, SUM (SubTask.StoryPoints) \
                FROM Issue \
                JOIN Project ON Issue.ProjectID = Project.ProjectID \
                JOIN SubTask ON Issue.IssueID = SubTask.IssueID \
                WHERE Project.Name = 'DAI DASy IDC5' \
                    AND Issue.Resolved BETWEEN " + "'" + cwStart + "'" + " AND " + "'" + cwEnd + "' \
                    AND Issue.SubTasks = 'None' \
                    AND (Issue.Assignee NOT LIKE '%Jan%' OR Issue.Assignee NOT LIKE '%Andrei%')"
        cursor.execute(query)
        sqlData = cursor.fetchall()
        dataset = pd.DataFrame(data=sqlData)
        dataset = dataset.fillna('')
        return dataset

    def getSubtasksStoryPointsOpen_IDC5(self, cursor, pd):
        query = "SELECT Project.Name, SUM (SubTask.StoryPoints) \
                FROM Issue \
                JOIN Project ON Issue.ProjectID = Project.ProjectID \
                JOIN SubTask ON Issue.IssueID = SubTask.IssueID \
                WHERE Project.Name = 'DAI DASy IDC5' \
                    AND Issue.Status <> 'Closed' \
                    AND SubTask.Status <> 'Closed' \
                    AND Issue.SubTasks <> 'None' \
                    AND (Issue.Assignee NOT LIKE '%Jan%' OR Issue.Assignee NOT LIKE '%Andrei%')"
        cursor.execute(query)
        sqlData = cursor.fetchall()
        dataset = pd.DataFrame(data=sqlData)
        dataset = dataset.fillna('')
        return dataset

    def getMainTasksStoryPointsOpen_IDC5(self, cursor, pd):
        query = "SELECT Project.Name, SUM (SubTask.StoryPoints) \
                FROM Issue \
                JOIN Project ON Issue.ProjectID = Project.ProjectID \
                JOIN SubTask ON Issue.IssueID = SubTask.IssueID \
                WHERE Project.Name = 'DAI DASy IDC5' \
                    AND Issue.Status <> 'Closed' \
                    AND SubTask.Status <> 'Closed' \
                    AND Issue.SubTasks = 'None' \
                    AND (Issue.Assignee NOT LIKE '%Jan%' OR Issue.Assignee NOT LIKE '%Andrei%')"
        cursor.execute(query)
        sqlData = cursor.fetchall()
        dataset = pd.DataFrame(data=sqlData)
        dataset = dataset.fillna('')
        return dataset


    def getMainTasksStoryPointsBlocked_IDC5(self, cursor, pd):
        query = "SELECT Project.Name, SUM(Issue.StoryPoints) \
                FROM Issue \
                JOIN Project ON Issue.ProjectID = Project.ProjectID \
                WHERE Project.Name = 'DAI DASy IDC5' \
                    AND Issue.Flagged LIKE '%Impediment%' \
                    AND Issue.Status != 'Closed' "
        cursor.execute(query)
        sqlData = cursor.fetchall()
        dataset = pd.DataFrame(data=sqlData)
        dataset = dataset.fillna('')
        return dataset
    
    # MAP ----------------------------------------------------------------------------
    def getSubtasksStoryPointsClosedInGivenCW_MAP(self, cursor, cwStart, cwEnd, pd):
        query = "SELECT Project.Name, SUM (SubTask.StoryPoints) \
                FROM Issue \
                JOIN Project ON Issue.ProjectID = Project.ProjectID \
                JOIN SubTask ON Issue.IssueID = SubTask.IssueID \
                WHERE Project.Name = 'DAI DASy MAP' \
                    AND SubTask.Resolved BETWEEN " + "'" + cwStart + "'" + " AND " + "'" + cwEnd + "' \
                    AND Issue.SubTasks <> 'None' \
                    AND (Issue.Assignee NOT LIKE '%Jan%' OR Issue.Assignee NOT LIKE '%Andrei%')"
        cursor.execute(query)
        sqlData = cursor.fetchall()
        dataset = pd.DataFrame(data=sqlData)
        dataset = dataset.fillna('')
        return dataset

    def getMainTasksStoryPointsClosedInGivenCW_MAP(self, cursor, cwStart, cwEnd, pd):
        query = "SELECT Project.Name, SUM (SubTask.StoryPoints) \
                FROM Issue \
                JOIN Project ON Issue.ProjectID = Project.ProjectID \
                JOIN SubTask ON Issue.IssueID = SubTask.IssueID \
                WHERE Project.Name = 'DAI DASy MAP' \
                    AND Issue.Resolved BETWEEN " + "'" + cwStart + "'" + " AND " + "'" + cwEnd + "' \
                    AND Issue.SubTasks = 'None' \
                    AND (Issue.Assignee NOT LIKE '%Jan%' OR Issue.Assignee NOT LIKE '%Andrei%')"
        cursor.execute(query)
        sqlData = cursor.fetchall()
        dataset = pd.DataFrame(data=sqlData)
        dataset = dataset.fillna('')
        return dataset

    def getSubtasksStoryPointsOpen_MAP(self, cursor, pd):
        query = "SELECT Project.Name, SUM (SubTask.StoryPoints) \
                FROM Issue \
                JOIN Project ON Issue.ProjectID = Project.ProjectID \
                JOIN SubTask ON Issue.IssueID = SubTask.IssueID \
                WHERE Project.Name = 'DAI DASy MAP' \
                    AND Issue.Status <> 'Closed' \
                    AND SubTask.Status <> 'Closed' \
                    AND Issue.SubTasks <> 'None' \
                    AND (Issue.Assignee NOT LIKE '%Jan%' OR Issue.Assignee NOT LIKE '%Andrei%')"
        cursor.execute(query)
        sqlData = cursor.fetchall()
        dataset = pd.DataFrame(data=sqlData)
        dataset = dataset.fillna('')
        return dataset

    def getMainTasksStoryPointsOpen_MAP(self, cursor, pd):
        query = "SELECT Project.Name, SUM (SubTask.StoryPoints) \
                FROM Issue \
                JOIN Project ON Issue.ProjectID = Project.ProjectID \
                JOIN SubTask ON Issue.IssueID = SubTask.IssueID \
                WHERE Project.Name = 'DAI DASy MAP' \
                    AND Issue.Status <> 'Closed' \
                    AND SubTask.Status <> 'Closed' \
                    AND Issue.SubTasks = 'None' \
                    AND (Issue.Assignee NOT LIKE '%Jan%' OR Issue.Assignee NOT LIKE '%Andrei%')"
        cursor.execute(query)
        sqlData = cursor.fetchall()
        dataset = pd.DataFrame(data=sqlData)
        dataset = dataset.fillna('')
        return dataset

    def getMainTasksStoryPointsBlocked_MAP(self, cursor, pd):
        query = "SELECT Project.Name, SUM(Issue.StoryPoints) \
                FROM Issue \
                JOIN Project ON Issue.ProjectID = Project.ProjectID \
                WHERE Project.Name = 'DAI DASy MAP' \
                    AND Issue.Flagged LIKE '%Impediment%' \
                    AND Issue.Status != 'Closed'"
        cursor.execute(query)
        sqlData = cursor.fetchall()
        dataset = pd.DataFrame(data=sqlData)
        dataset = dataset.fillna('')
        return dataset


    # JLR ----------------------------------------------------------------------------
    def getSubtasksStoryPointsClosedInGivenCW_JLR(self, cursor, cwStart, cwEnd, pd):
        query = "SELECT Project.Name, SUM (SubTask.StoryPoints) \
                FROM Issue \
                JOIN Project ON Issue.ProjectID = Project.ProjectID \
                JOIN SubTask ON Issue.IssueID = SubTask.IssueID \
                WHERE Project.Name = 'DASy JLR L663' \
                    AND SubTask.Resolved BETWEEN " + "'" + cwStart + "'" + " AND " + "'" + cwEnd + "' \
                    AND Issue.SubTasks <> 'None' \
                    AND (Issue.Assignee NOT LIKE '%Jan%' OR Issue.Assignee NOT LIKE '%Andrei%')"
        cursor.execute(query)
        sqlData = cursor.fetchall()
        dataset = pd.DataFrame(data=sqlData)
        dataset = dataset.fillna('')
        return dataset

    def getMainTasksStoryPointsClosedInGivenCW_JLR(self, cursor, cwStart, cwEnd, pd):
        query = "SELECT Project.Name, SUM (SubTask.StoryPoints) \
                FROM Issue \
                JOIN Project ON Issue.ProjectID = Project.ProjectID \
                JOIN SubTask ON Issue.IssueID = SubTask.IssueID \
                WHERE Project.Name = 'DASy JLR L663' \
                    AND Issue.Resolved BETWEEN " + "'" + cwStart + "'" + " AND " + "'" + cwEnd + "' \
                     AND Issue.SubTasks = 'None' \
                    AND (Issue.Assignee NOT LIKE '%Jan%' OR Issue.Assignee NOT LIKE '%Andrei%')"
        cursor.execute(query)
        sqlData = cursor.fetchall()
        dataset = pd.DataFrame(data=sqlData)
        dataset = dataset.fillna('')
        return dataset

    def getSubtasksStoryPointsOpen_JLR(self, cursor, pd):
        query = "SELECT Project.Name, SUM (SubTask.StoryPoints) \
                FROM Issue \
                JOIN Project ON Issue.ProjectID = Project.ProjectID \
                JOIN SubTask ON Issue.IssueID = SubTask.IssueID \
                WHERE Project.Name = 'DASy JLR L663' \
                    AND Issue.Status <> 'Closed' \
                    AND SubTask.Status <> 'Closed' \
                    AND Issue.SubTasks <> 'None' \
                    AND (Issue.Assignee NOT LIKE '%Jan%' OR Issue.Assignee NOT LIKE '%Andrei%')"
        cursor.execute(query)
        sqlData = cursor.fetchall()
        dataset = pd.DataFrame(data=sqlData)
        dataset = dataset.fillna('')
        return dataset

    def getMainTasksStoryPointsOpen_JLR(self, cursor, pd):
        query = "SELECT Project.Name, SUM (SubTask.StoryPoints) \
                FROM Issue \
                JOIN Project ON Issue.ProjectID = Project.ProjectID \
                JOIN SubTask ON Issue.IssueID = SubTask.IssueID \
                WHERE Project.Name = 'DASy JLR L663' \
                    AND Issue.Status <> 'Closed' \
                    AND SubTask.Status <> 'Closed' \
                    AND Issue.SubTasks = 'None' \
                    AND (Issue.Assignee NOT LIKE '%Jan%' OR Issue.Assignee NOT LIKE '%Andrei%')"
        cursor.execute(query)
        sqlData = cursor.fetchall()
        dataset = pd.DataFrame(data=sqlData)
        dataset = dataset.fillna('')
        return dataset

    def getMainTasksStoryPointsBlocked_JLR(self, cursor, pd):
        query = "SELECT Project.Name, SUM(Issue.StoryPoints) \
                FROM Issue \
                JOIN Project ON Issue.ProjectID = Project.ProjectID \
                WHERE Project.Name = 'DASy JLR L663' \
                    AND Issue.Flagged LIKE '%Impediment%' \
                    AND Issue.Status != 'Closed' "
        cursor.execute(query)
        sqlData = cursor.fetchall()
        dataset = pd.DataFrame(data=sqlData)
        dataset = dataset.fillna('')
        return dataset

    def insertAllResults(self, sqliteDb, cursor, resultDictOverall, resultDictIDC5, resultDictMAP, resultDictJLR, logger):
        overallTable = 'WeeklyReportResults'
        idc5Table = 'ResultsIDC5'
        mapTable = 'ResultsMAP'
        jlrTable = 'ResultsJLR'

        self.insertResultsPerProject(cursor, resultDictOverall, overallTable)
        resultDictIDC5['ResultsID'] = resultDictMAP['ResultsID'] = resultDictJLR['ResultsID'] = cursor.lastrowid

        self.insertResultsPerProject(cursor, resultDictIDC5, idc5Table)
        self.insertResultsPerProject(cursor, resultDictMAP, mapTable)
        self.insertResultsPerProject(cursor, resultDictJLR, jlrTable)

        sqliteDb.commit()
        logger.info("All results has been written in database")
        return

    def insertResultsPerProject(self, cursor, resultDict, tableName):
        sql = 'INSERT INTO {} ({}) VALUES ({})'.format(tableName, ','.join(resultDict),	','.join(['?']*len(resultDict)))
        cursor.execute(sql, list(resultDict.values()))
        return

    def updateAllResults(self, sqliteDb, cursor, resultDictOverall, resultDictIDC5, resultDictMAP, resultDictJLR, logger):
        overallTable = 'WeeklyReportResults'
        idc5Table = 'ResultsIDC5'
        mapTable = 'ResultsMAP'
        jlrTable = 'ResultsJLR'

        self.updateResultsPerProject(cursor, resultDictOverall, overallTable)
        self.updateResultsPerProject(cursor, resultDictIDC5, idc5Table)
        self.updateResultsPerProject(cursor, resultDictMAP, mapTable)
        self.updateResultsPerProject(cursor, resultDictJLR, jlrTable)

        sqliteDb.commit()
        logger.info("All results has been written in database")
        return

    def updateResultsPerProject(self, cursor, resultDict, tableName):
        newValues = []
        for key, value in resultDict.items():
            newPair = "{} = "'"{}"'"".format(key, value)
            newValues.append(newPair)
            columns_newValues = ', '.join(newValues)
        sql = '''UPDATE {} SET {} WHERE ResultsID={}'''.format(tableName, columns_newValues, str(resultDict['ResultsID']))
        cursor.execute(sql, resultDict)
        return

class ExcelOperations():

    def __new__(cls):
        self = object.__new__(cls)
        return self

    def openGivenWorkbook(self, filePath):
        self.workbook = openpyxl.load_workbook(filePath)
        return self.workbook

    def openNewWorkbook(self):
        self.workbook = openpyxl.Workbook()
        return self.workbook

    def openGivenWorksheet(self, workbook, worksheet):
        self.worksheet = workbook[worksheet]
        return self.worksheet
        
    def openGivenWorkbookWithActiveWorksheet(self, filePath):
        self.workbook = openpyxl.load_workbook(filePath)
        self.mainWorksheet = self.workbook.active
        return self.mainWorksheet

    def getAllDataRange(self, worksheet):
        self.fullRange = "A1:" + get_column_letter(worksheet.max_column) + str(worksheet.max_row)
        return self.fullRange

    def setOverallFilter(self, worksheet, dataRange):
        worksheet.auto_filter.ref = dataRange
        return

    def deleteRows(self, worksheet, rowIndex, rowsAmount):
        self.remainder = worksheet.gutter(rowIndex, rowsAmount, worksheet.max_column)
        worksheet.move_cells(min_col = rowIndex + rowsAmount, offset=-rowsAmount, row_or_col="column")
        for col in self.remainder:
            for row in range(worksheet.min_row, worksheet.max_row+1):
                if (row, col) in worksheet._cells:
                    del worksheet._cells[row, col]

        return

    def saveWorkbook(self, workbook, path, filename, extension):
        workbook.save(path + filename + extension)
        return

    def writeDict(self, sheet, next_row, dict, noOfItemsToExclude):
        dictToWriteLength = len(dict) - noOfItemsToExclude
        for i in range(dictToWriteLength):
            resultOverall = list(dict.items())[i]
            sheet.cell(column=1, row=next_row, value=resultOverall[0])
            sheet.cell(column=2, row=next_row, value=resultOverall[1])
            next_row += 1

        return sheet, next_row

    def writeResults(self, newExcelWb, sheet, initStruct, resultDictOverall, resultDictIDC5, resultDictMAP, resultDictJLR, logger):
         # header
        sheet.cell(column=1, row=1, value=initStruct['easy_week_number'])
        sheet.cell(column=2, row=1, value='Week start date:')
        sheet.cell(column=3, row=1, value=initStruct['week_start_date'].strftime('%d-%m-%Y'))
        sheet.cell(column=4, row=1, value='Week end date:')
        sheet.cell(column=5, row=1, value=initStruct['week_end_date'].strftime('%d-%m-%Y'))
        sheet.cell(column=1, row=2, value='Data source file:')
        sheet.cell(column=2, row=2, value=initStruct['last_jira_export_file_path'])

        # results per project
        next_row=4
        sheet.cell(column=1, row=next_row, value='Project: ')
        sheet.cell(column=2, row=next_row, value=resultDictOverall['ProjectName'])
        next_row += 1
        response = self.writeDict(sheet, next_row, resultDictOverall, 7)
        sheet = response[0]
        next_row = response[1]

        next_row += 1
        sheet.cell(column=1, row=next_row, value='Project: ')
        sheet.cell(column=2, row=next_row, value=resultDictIDC5['ProjectName'])
        next_row += 1
        response = self.writeDict(sheet, next_row, resultDictIDC5, 3)
        sheet = response[0]
        next_row = response[1]

        next_row += 1
        sheet.cell(column=1, row=next_row, value='Project: ')
        sheet.cell(column=2, row=next_row, value=resultDictMAP['ProjectName'])
        next_row += 1
        response = self.writeDict(sheet, next_row, resultDictMAP, 3)
        sheet = response[0]
        next_row = response[1]

        next_row += 1
        sheet.cell(column=1, row=next_row, value='Project: ')
        sheet.cell(column=2, row=next_row, value=resultDictJLR['ProjectName'])
        next_row += 1
        response = self.writeDict(sheet, next_row, resultDictJLR, 3)
        sheet = response[0]
        next_row = response[1]

        # formatting columns width
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = sheet.column_dimensions['C'].width = sheet.column_dimensions['D'].width = sheet.column_dimensions['E'].width = 15

        resultsExcelFile = pjoin(initStruct['data_destinations_folder'],'Results_excel','EAD2_DASy_TestCenter_Alten_WeeklyReporting_2019_' + initStruct['easy_week_number'] + '.xlsx')
        newExcelWb.save(resultsExcelFile)
        newExcelWb.close()

        return resultsExcelFile

class FileOperations():

    def __new__(cls):
        self = object.__new__(cls)
        return self

    def getLastExportedFileByNamePattern(self, pathToFiles, fileNamePattern):
        filesList = []
        for s in os.listdir(pathToFiles):
            filesList.append(s)
        filesList = fnmatch.filter(filesList, fileNamePattern)
        if os.path.isfile(os.path.join(pathToFiles, s)):
            filesList.sort(key=lambda s: os.path.getctime(os.path.join(pathToFiles, s)))

        fileName = filesList.pop()
        return fileName

    def checkFileExistence(self, pathToFile):
        self.fileExists = os.path.exists(pathToFile)
        return self.fileExists

    def removeFile(self, logger, exportSavePath, fullFilename):
        if os.path.isfile(exportSavePath):
            os.remove(exportSavePath)
            logger.info('Done')
        else:
            logger.info("A previous {} file was not found. Continue with saving the new one.".format(fullFilename))

        return

    def saveFile(self, logger, exportSavePath, responseContent):
        #file.write(responseContent)
        #file.close()
        logger.info('Done')

        return

class UrlOperations():

    def __new__(cls):
        self = object.__new__(cls)
        return self

    def checkServerResponse(self, url):
        #response = requests.get(url, auth=HTTPDigestAuth(configStruct['username'], configStruct['password']), allow_redirects=True, stream=True)
        response = 200
        return response

class ReadInputs(object, metaclass=Singleton):
    def ParseIniFile(self, initStruct, fileOps):
        # 1. check ini file existence in working folder
        workingFolderPath = os.path.dirname(sys.argv[0])
        fileExists = fileOps.checkFileExistence(workingFolderPath)
        if fileExists == False:
            return False
        else:
            initStruct['working_folder'] = os.path.dirname(sys.argv[0])
            initStruct['iniFileName'] = INIFILENAME
            # 2. read the ini file
            init = configparser.ConfigParser(allow_no_value=True)
            iniFilePath = pjoin(initStruct['working_folder'], initStruct['iniFileName'])
            init.read(iniFilePath)
            # from Generic section
            initStruct['week_number'] = int(init.get('generic', 'week_number'))
            initStruct['datetime_week_number'] = "2019-W" + str(initStruct['week_number'] - 1)
            initStruct['easy_week_number'] = "CW" + str(initStruct['week_number'])
            initStruct['mapping_file'] = init.get('generic', 'mapping_file')
            initStruct['logFilename'] = init.get('generic', 'logFilename')
            initStruct['logToConsole'] = init.get('generic', 'logToConsole')
            initStruct['logToFile'] = init.get('generic', 'logToFile')
            # from data_sources section
            initStruct['data_sources_folder'] = init.get('data_sources', 'data_sources_folder')
            initStruct['jira_export_folder'] = init.get('data_sources', 'jira_export_folder')
            initStruct['jira_export_filename_pattern'] = init.get('data_sources', 'jira_export_filename_pattern')
            initStruct['jira_export_file_extension'] = init.get('data_sources', 'jira_export_file_extension')
            initStruct['jira_url_domain'] = init.get('data_sources', 'jira_url_domain')
            # from data_destinations section
            initStruct['data_destinations_folder'] = init.get('data_destinations', 'data_destinations_folder')
            initStruct['sqlite_db_filename'] = init.get('data_destinations', 'sqlite_db_filename')
            return True

    def GetUserCredentials(self, logger, initStruct):
        initStruct['username'] = getpass.getuser()
        initStruct['password'] = getpass.getpass('Input password for {} user: '.format(initStruct['username']))
        initStruct['token'] = input('Input access token: ')
        if initStruct['password']:
            if initStruct['token']:
                return
        else:
            logger.info('Password or token are empty. Try again.')
            self.GetUserCredentials(logger,initStruct)

class Main(object, metaclass=Singleton):
    def CheckExternalAvailability(self, fileOps, web, initStruct, logger):
            mappingfileExists = fileOps.checkFileExistence(initStruct['mapping_file_path'])
            if mappingfileExists == False:
                logger.info('The Column Mapping file is not present in the specified folder or with proper name or extension. Check it or create it and start again.')
                sys.exit()

            serverResponseCode = web.checkServerResponse(initStruct['jira_url_domain'])
            if serverResponseCode != 200:
                logger.info("Server responded with error: {0}. Check the connection to Jira server.".format(str(serverResponseCode)))

            jiraExportfileExists = fileOps.checkFileExistence(initStruct['last_jira_export_file_path'])
            if jiraExportfileExists == False:
                logger.info('Excel file with exported results from Jira is not present in the specified folder or with proper name or extension. Check it or create it and start again.')
                sys.exit()

            databaseFileExists = fileOps.checkFileExistence(initStruct['sqlite_db_path'])
            if databaseFileExists == False:
                logger.info('Database file is not present in the specified folder. Check location, file name or Ini file setting. This script will not continue without it.')
                sys.exit()
            return

    def ReadDataThenWrite(self, sqliteDb, cursor, allRows, issueMap, excelColIdxDict, activeWorksheet, issueDataDict, subtaskDataDict, dbProjectData, logger):
        for i in range(2, allRows + 1):
            loggerText1 = "read from source..."
            for key, value in issueMap.items():
                strDbKey = str(key)
                strExcelKey = str(value)
                colIndex = int(excelColIdxDict[strExcelKey])
                rowColValue = activeWorksheet.cell(i, colIndex).value
                dataType = type(rowColValue)
                if dataType == str:
                    issueDataDict[strDbKey] = rowColValue.encode('utf-8')
                elif dataType == datetime:
                    # in SQLite if you use the TEXT storage class to store date and time value, you need to use the ISO8601 string format 
                    issueDataDict[strDbKey] = datetime.strftime(rowColValue, '%y-%m-%d %H:%M')
                else:
                    issueDataDict[strDbKey] = str(rowColValue)

            # step 5 - adapt data
            issueDataDict["IssueType"] = issueDataDict["IssueType"].decode()
            issueDataDict["ProjectID"] = issueDataDict["ProjectID"].decode()
            issueDataDict["Summary"] = issueDataDict["Summary"].decode()
            issueDataDict["IssueKey"] = issueDataDict["IssueKey"].decode()
            issueDataDict["Assignee"] = issueDataDict["Assignee"].decode()
            issueDataDict["Priority"] = issueDataDict["Priority"].decode()
            issueDataDict["Status"] = issueDataDict["Status"].decode()
            

            for row in dbProjectData:
                if issueDataDict["ProjectID"] == row[1]:
                    issueDataDict["ProjectID"] = row[0]
            
            if issueDataDict["IssueType"] != "Sub-task":
                loggerText2 = ""
                # check if Issue already exists
                cursor.execute('SELECT * FROM issue WHERE IssueKey=?', (issueDataDict["IssueKey"],))
                existingIssue = cursor.fetchall()
                if not existingIssue:
                    # insert row
                    sql = 'INSERT INTO issue ({}) VALUES ({})'.format(','.join(issueDataDict),	','.join(['?']*len(issueDataDict)))
                    loggerText1 = "insert issue in database..."
                    cursor.execute(sql, list(issueDataDict.values()))
                else:
                    issueId = existingIssue[0][0]
                    # update row
                    newValues = []
                    for key, value in issueDataDict.items():
                        if type(value) is str: 
                            myString = str(value)
                            value = myString.strip("\"").replace("\"", "")
                        newPair = "{} = "'"{}"'"".format(key, value)
                        newValues.append(newPair)
                    columns_newValues = ', '.join(newValues)
                    sql = 'UPDATE issue SET {} WHERE IssueID={}'.format(columns_newValues, str(issueId))
                    loggerText1 = "update issue in database..."
                    cursor.execute(sql)
            else:
                loggerText1 = ""
                subtaskDataDict = issueDataDict.copy()
                summaryList = issueDataDict["Summary"].split(" ", 1)
                getParentTask = summaryList[0]
                cursor.execute('SELECT * FROM issue WHERE IssueKey=?', (getParentTask,))
                parent_task_rows = cursor.fetchall()
                if not parent_task_rows:
                    subtaskDataDict["IssueID"] = 999999
                else:
                    subtaskDataDict["IssueID"] = parent_task_rows[0][0]

                subtaskDataDict["Classification"] = ""
                subtaskDataDict["Factor"] = ""
                del subtaskDataDict["ProjectID"]
                del subtaskDataDict["SubTasks"]
                del subtaskDataDict["ProblemType"]
                del subtaskDataDict["IssueType"]

                # check if Subtask already exists
                cursor.execute('SELECT * FROM subtask WHERE IssueKey=?', (subtaskDataDict["IssueKey"],))
                existingSubtask = cursor.fetchall()
                if not existingSubtask:
                    # insert row
                    sql = 'INSERT INTO subtask ({}) VALUES ({})'.format(
                        ','.join(subtaskDataDict),
                        ','.join(['?']*len(subtaskDataDict)))
                    loggerText2 = "insert subtask in database..."
                    cursor.execute(sql, list(subtaskDataDict.values()))
                else:
                    subtaskId = existingSubtask[0][0]
                    # update row
                    newValues = []
                    for key, value in subtaskDataDict.items():
                        if type(value) is str: 
                            myString = str(value)
                            value = myString.strip("\"").replace("\"", "")
                        newPair = "{} = "'"{}"'"".format(key, value)
                        newValues.append(newPair)
                    columns_newValues = ', '.join(newValues)
                    sql = '''UPDATE subtask SET {} WHERE IssueID={}'''.format(columns_newValues, str(subtaskId))
                    loggerText2 = "update subtask in database..."
                    cursor.execute(sql)

            sqliteDb.commit()
            logger.info("Row {0}: {1} {2} done.".format(i, loggerText1, loggerText2))

    def GetResultsForSlide5(self, dbOps, cursor, initStruct, resultDictOverall, resultDictIDC5, resultDictJLR, resultDictMAP, logger):
        try:
            # get created Issues (not subtasks) in specific CW for all projects
            dataset = dbOps.getMainTasksCreatedInGivenCW(cursor, initStruct['week_start_date'].strftime('%y-%m-%d %H:%M'), initStruct['week_end_date'].strftime('%y-%m-%d %H:%M'), pd)
            #logger.info("    Open tickets:")

            tasksOpen = dataset[dataset [4] == "Open"]   # take care to Status column index in dataset if add/remove columns in SELECT
            resultDictOverall['OpenTicketsTotal'] = tasksOpen.shape[0]
            #logger.info('\tOverall: {}'.format(resultsList.resultDictOverall['OpenTicketsTotal']))

            tasksOpenIDC5 = tasksOpen[tasksOpen[0] == 'DAI DASy IDC5']   # take care to Project Name column index in dataset if add/remove columns in SELECT
            resultDictIDC5['OpenTicketsTotal'] = tasksOpenIDC5.shape[0]
            #logger.info('\tIDC5: {}'.format(resultsList.resultDictIDC5['OpenTicketsTotal']))

            tasksOpenJLR = tasksOpen[tasksOpen[0] == 'DASy JLR L663']   # take care to Project Name column index in dataset if add/remove columns in SELECT
            resultDictJLR['OpenTicketsTotal'] = tasksOpenJLR.shape[0]
            #logger.info('\tJLR: {}'.format(resultsList.resultDictJLR['OpenTicketsTotal']))

            tasksOpenMAP = tasksOpen[tasksOpen[0] == 'DAI DASy MAP']   # take care to Project Name column index in dataset if add/remove columns in SELECT
            resultDictMAP['OpenTicketsTotal'] = tasksOpenMAP.shape[0]
            #logger.info('\tMAP: {}'.format(resultsList.resultDictMAP['OpenTicketsTotal']))

            dataset = dbOps.getMainTasksInProgress(cursor, pd)
            #logger.info("    In Progress tickets:")
            tasksInProgress = dataset[dataset[4] == "In Progress"]
            resultDictOverall['InProgressTicketsTotal'] = dataset.shape[0]
            #logger.info('\tOverall: {}'.format(resultsList.resultDictOverall['InProgressTicketsTotal']))

            tasksInProgressIDC5 = tasksInProgress[tasksInProgress[0] == 'DAI DASy IDC5'] & tasksInProgress[tasksInProgress[0] == 'DASy Platform Proj']
            resultDictIDC5['InProgressTicketsTotal'] = tasksInProgressIDC5.shape[0]
            #logger.info('\tIDC5: {}'.format(resultsList.resultDictIDC5['InProgressTicketsTotal']))

            tasksInProgressJLR = tasksInProgress[tasksInProgress[0] == 'DASy JLR L663']
            resultDictJLR['InProgressTicketsTotal'] = tasksInProgressJLR.shape[0]
            #logger.info('\tJLR: {}'.format(resultsList.resultDictJLR['InProgressTicketsTotal']))

            tasksInProgressMAP = tasksInProgress[tasksInProgress[0] == 'DAI DASy MAP']
            resultDictMAP['InProgressTicketsTotal'] = tasksInProgressMAP.shape[0]
            #logger.info('\tMAP: {}'.format(resultsList.resultDictMAP['InProgressTicketsTotal']))

            dataset = dbOps.getMainTasksClosedInGivenCW(cursor, initStruct['week_start_date'].strftime('%y-%m-%d %H:%M'), initStruct['week_end_date'].strftime('%y-%m-%d %H:%M'), pd)
            #logger.info("    Closed tickets:")

            tasksClosed = dataset[dataset[4] == "Closed"]
            resultDictOverall['ClosedTicketsTotal_perCW'] = tasksClosed.shape[0]
            #logger.info('\tOverall: {}'.format(resultsList.resultDictOverall['ClosedTicketsTotal_perCW']))

            tasksClosedIDC5 = tasksClosed[tasksClosed[0] == 'DAI DASy IDC5']   # take care to Project Name column index in dataset if add/remove columns in SELECT
            resultDictIDC5['ClosedTicketsTotal_perCW'] = tasksClosedIDC5.shape[0]
            #logger.info('\tIDC5: {}'.format(resultsList.resultDictIDC5['ClosedTicketsTotal_perCW']))

            tasksClosedJLR = tasksClosed[tasksClosed[0] == 'DASy JLR L663']   # take care to Project Name column index in dataset if add/remove columns in SELECT
            resultDictJLR['ClosedTicketsTotal_perCW'] = tasksClosedJLR.shape[0]
            #logger.info('\tJLR: {}'.format(resultsList.resultDictJLR['ClosedTicketsTotal_perCW']))

            tasksClosedMAP = tasksClosed[tasksClosed[0] == 'DAI DASy MAP']   # take care to Project Name column index in dataset if add/remove columns in SELECT
            resultDictMAP['ClosedTicketsTotal_perCW'] = tasksClosedMAP.shape[0]
            #logger.info('\tMAP: {}'.format(resultsList.resultDictMAP['ClosedTicketsTotal_perCW']))
        
        except KeyError:
            exit()

    def GetResultsForSlide6(self, dbOps, cursor, initStruct, resultDictIDC5, resultDictJLR, resultDictMAP, logger):
        
        try:
            dataset = dbOps.getSubtasksStoryPointsCreatedInGivenCW(cursor, initStruct['week_start_date'].strftime('%y-%m-%d %H:%M'), initStruct['week_end_date'].strftime('%y-%m-%d %H:%M'), pd)

            #logger.info("    JLR: ")
            totalTicketsJLR = dataset[dataset[0] == 'DASy JLR L663']    # take care to Project Name column index in dataset if add/remove columns in SELECT
            resultDictJLR['OpenTickets_perCW_Count'] = totalTicketsJLR.shape[0]
            resultDictJLR['OpenTickets_perCW_SPsum'] = float(totalTicketsJLR[5].sum(skipna = True))
            #logger.info('\tNew tickets: {}'.format(resultsList.resultDictJLR['OpenTickets_perCW_Count']))
            #logger.info('\tSubTasks story points sum: {}'.format(resultsList.resultDictJLR['OpenTickets_perCW_SPsum']))

            #logger.info("    IDC5: ")
            totalTicketsIDC5 = dataset[dataset[0] == 'DAI DASy IDC5']   # take care to Project Name column index in dataset if add/remove columns in SELECT
            resultDictIDC5['OpenTickets_perCW_Count'] = totalTicketsIDC5.shape[0]
            resultDictIDC5['OpenTickets_perCW_SPsum'] = float(totalTicketsIDC5[5].sum(skipna = True))
            #logger.info('\tNew tickets: {}'.format(resultsList.resultDictIDC5['OpenTickets_perCW_Count']))
            #logger.info('\tSubTasks story points sum: {}'.format(resultsList.resultDictIDC5['OpenTickets_perCW_SPsum']))

            #logger.info("    MAP: ")
            totalTicketsMAP = dataset[dataset[0] == 'DAI DASy MAP']    # take care to Project Name column index in dataset if add/remove columns in SELECT
            resultDictMAP['OpenTickets_perCW_Count'] = totalTicketsMAP.shape[0]
            resultDictMAP['OpenTickets_perCW_SPsum'] = float(totalTicketsMAP[5].sum(skipna = True))
            #logger.info('\tNew tickets: {}'.format(resultsList.resultDictMAP['OpenTickets_perCW_Count']))
            #logger.info('\tSubTasks story points sum: {}'.format(resultsList.resultDictMAP['OpenTickets_perCW_SPsum']))

        except KeyError:
            print("KeyError: Please, insert subtasks")
            exit()

    def GetResultsForSlide7(self, dbOps, cursor, initStruct, resultDictOverall, logger):

        try:

            # get SP of Closed Subtasks of tasks (with subtasks)
            dataset1 = dbOps.getSubtasksStoryPointsClosedInGivenCW(cursor, initStruct['week_start_date'].strftime('%y-%m-%d %H:%M'), initStruct['week_end_date'].strftime('%y-%m-%d %H:%M'), pd)
            if dataset1[1].iloc[0] != '':
                totalSPofClosedSubtasks = dataset1[1].iloc[0]
            else:
                totalSPofClosedSubtasks = 0

            # get SP of Closed tasks (without subtasks)
            dataset2 = dbOps.getMainTasksStoryPointsClosedInGivenCW(cursor, initStruct['week_start_date'].strftime('%y-%m-%d %H:%M'), initStruct['week_end_date'].strftime('%y-%m-%d %H:%M'), pd)
            if dataset2[1].iloc[0] != '':
                totalSPofClosedTasks = dataset1[1].iloc[0]
            else:
                totalSPofClosedTasks = 0

            resultDictOverall['ActualTestedReq_SPsum'] = float(totalSPofClosedSubtasks + totalSPofClosedTasks)
            #logger.info("    Actual tested Req. (IST): {}".format(resultDictOverall['ActualTestedReq_SPsum']))

            # get SP of Open Subtasks of tasks (with subtasks)
            dataset3 = dbOps.getSubtasksStoryPointsOpen(cursor, pd)
            if dataset3[1].iloc[0] != '':
                totalSPofOpenSubtasks = dataset3[1].iloc[0]
            else:
                totalSPofOpenSubtasks = 0

            # get SP of Open tasks (without subtasks)
            dataset4 = dbOps.getMainTasksStoryPointsOpen(cursor, pd)
            if dataset4[1].iloc[0] != '':
                totalSPofOpenTasks = dataset4[1].iloc[0]
            else:
                totalSPofOpenTasks = 0

            #get SP of Blocked tasks (without subtasks)
            dataset5 = dbOps.getMainTasksStoryPointsBlocked(cursor, pd)
            if dataset5[1].iloc[0] != '':
                totalSPofBlockedTasks = dataset5[1].iloc[0]
            else:
                totalSPofBlockedTasks = 0

            resultDictOverall['ActualInProgress_SPsum'] = float(totalSPofOpenSubtasks + totalSPofOpenTasks)
            #logger.info("    Actual In Progress (IST): {}".format(resultDictOverall['ActualInProgress_SPsum']))
            resultDictOverall['ActualBlockedReq_SPsum'] = float(totalSPofBlockedTasks)
            #logger.info("    Actual blocked Req. (IST): {}".format(resultDictOverall['ActualBlockedReq_SPsum']))
            resultDictOverall['ActualSummaryReq_SPsum'] = float(resultDictOverall['ActualTestedReq_SPsum'] + resultDictOverall['ActualInProgress_SPsum'] + resultDictOverall['ActualBlockedReq_SPsum'])
            #logger.info("    Actual Summary (IST): {}".format(resultDictOverall['ActualSummaryReq_SPsum']))

        except KeyError:
            print("KeyError: Please, insert subtasks")
            exit()

    def GetResultsForSlide8(self, dbOps, cursor, initStruct, resultDictIDC5, logger):
        
        try:
            # get SP of Closed Subtasks of tasks (with subtasks) for IDC5
            dataset1 = dbOps.getSubtasksStoryPointsClosedInGivenCW_IDC5(cursor, initStruct['week_start_date'].strftime('%y-%m-%d %H:%M'), initStruct['week_end_date'].strftime('%y-%m-%d %H:%M'), pd)
            if dataset1[1].iloc[0] != '':
                totalSPofClosedSubtasksIDC5 = dataset1[1].iloc[0]
            else:
                totalSPofClosedSubtasksIDC5 = 0

            # get SP of Closed tasks (without subtasks) for IDC5
            dataset2 = dbOps.getMainTasksStoryPointsClosedInGivenCW_IDC5(cursor, initStruct['week_start_date'].strftime('%y-%m-%d %H:%M'), initStruct['week_end_date'].strftime('%y-%m-%d %H:%M'), pd)
            if dataset2[1].iloc[0] != '':
                totalSPofClosedTasksIDC5 = dataset1[1].iloc[0]
            else:
                totalSPofClosedTasksIDC5 = 0

            resultDictIDC5['ActualTestedReq_SPsum'] = float(totalSPofClosedSubtasksIDC5 + totalSPofClosedTasksIDC5)
            #logger.info("    Actual tested Req. (IST): {}".format(resultDictIDC5['ActualTestedReq_SPsum']))

            # get SP of Open Subtasks of tasks (with subtasks) for IDC5
            dataset3 = dbOps.getSubtasksStoryPointsOpen_IDC5(cursor, pd)
            if dataset3[1].iloc[0] != '':
                totalSPofOpenSubtasksIDC5 = dataset3[1].iloc[0]
            else:
                totalSPofOpenSubtasksIDC5 = 0

            # get SP of Open tasks (without subtasks) for IDC5
            dataset4 = dbOps.getMainTasksStoryPointsOpen_IDC5(cursor, pd)
            if dataset4[1].iloc[0] != '':
                totalSPofOpenTasksIDC5 = dataset4[1].iloc[0]
            else:
                totalSPofOpenTasksIDC5 = 0

            #get SP of Blocked tasks (without subtasks) for IDC5
            dataset5 = dbOps.getMainTasksStoryPointsBlocked_IDC5(cursor, pd)
            if dataset5[1].iloc[0] != '':
                totalSPofBlockedTasksIDC5 = dataset5[1].iloc[0]
            else:
                totalSPofBlockedTasks = 0

            resultDictIDC5['ActualInProgress_SPsum'] = float(totalSPofOpenSubtasksIDC5 + totalSPofOpenTasksIDC5)
            #logger.info("    Actual In Progress (IST): {}".format(resultDictIDC5['ActualInProgress_SPsum']))
            resultDictIDC5['ActualBlockedReq_SPsum'] = float(totalSPofBlockedTasksIDC5)
            #logger.info("    Actual blocked Req. (IST): {}".format(resultDictIDC5['ActualBlockedReq_SPsum']))
            resultDictIDC5['ActualSummaryReq_SPsum'] = float(resultDictIDC5['ActualTestedReq_SPsum'] + resultDictIDC5['ActualInProgress_SPsum'] + resultDictIDC5['ActualBlockedReq_SPsum'])
            #logger.info("    Actual Summary (IST): {}".format(resultDictIDC5['ActualSummaryReq_SPsum']))

        except KeyError:
            print("KeyError: Please, insert subtasks")
            exit()

    def GetResultsForSlide9(self, dbOps, cursor, initStruct, resultDictMAP, logger):

        try:
            # get SP of Closed Subtasks of tasks (with subtasks) for MAP
            dataset1 = dbOps.getSubtasksStoryPointsClosedInGivenCW_MAP(cursor, initStruct['week_start_date'].strftime('%y-%m-%d %H:%M'), initStruct['week_end_date'].strftime('%y-%m-%d %H:%M'), pd)
            if dataset1[1].iloc[0] != '':
                totalSPofClosedSubtasksMAP = dataset1[1].iloc[0]
            else:
                totalSPofClosedSubtasksMAP = 0

            # get SP of Closed tasks (without subtasks) for MAP
            dataset2 = dbOps.getMainTasksStoryPointsClosedInGivenCW_MAP(cursor, initStruct['week_start_date'].strftime('%y-%m-%d %H:%M'), initStruct['week_end_date'].strftime('%y-%m-%d %H:%M'), pd)
            if dataset2[1].iloc[0] != '':
                totalSPofClosedTasksMAP = dataset1[1].iloc[0]
            else:
                totalSPofClosedTasksMAP = 0

            resultDictMAP['ActualTestedReq_SPsum'] = float(totalSPofClosedSubtasksMAP + totalSPofClosedTasksMAP)
            #logger.info("    Actual tested Req. (IST): {}".format(resultDictMAP['ActualTestedReq_SPsum']))

            # get SP of Open Subtasks of tasks (with subtasks) for MAP
            dataset3 = dbOps.getSubtasksStoryPointsOpen_MAP(cursor, pd)
            if dataset3[1].iloc[0] != '':
                totalSPofOpenSubtasksMAP = dataset3[1].iloc[0]
            else:
                totalSPofOpenSubtasksMAP = 0

            # get SP of Open tasks (without subtasks) for MAP
            dataset4 = dbOps.getMainTasksStoryPointsOpen_MAP(cursor, pd)
            if dataset4[1].iloc[0] != '':
                totalSPofOpenTasksMAP = dataset4[1].iloc[0]
            else:
                totalSPofOpenTasksMAP = 0

            #get SP of Blocked tasks (without subtasks) for MAP
            dataset5 = dbOps.getMainTasksStoryPointsBlocked_MAP(cursor, pd)
            if dataset5[1].iloc[0] != '':
                totalSPofBlockedTasksMAP = dataset5[1].iloc[0]
            else:
                totalSPofBlockedTasksMAP = 0

            resultDictMAP['ActualInProgress_SPsum'] = float(totalSPofOpenSubtasksMAP + totalSPofOpenTasksMAP)
            #logger.info("    Actual In Progress (IST): {}".format(resultDictMAP['ActualInProgress_SPsum']))
            resultDictMAP['ActualBlockedReq_SPsum'] = float(totalSPofBlockedTasksMAP)
            #logger.info("    Actual blocked Req. (IST): {}".format(resultDictMAP['ActualBlockedReq_SPsum']))
            resultDictMAP['ActualSummaryReq_SPsum'] = float(resultDictMAP['ActualTestedReq_SPsum'] + resultDictMAP['ActualInProgress_SPsum'] + resultDictMAP['ActualBlockedReq_SPsum'])
            #logger.info("    Actual Summary (IST): {}".format(resultDictMAP['ActualSummaryReq_SPsum']))

        except KeyError:
            print("KeyError: Please, insert subtasks")
            exit()

    def GetResultsForSlide10(self, dbOps, cursor, initStruct, resultDictJLR, logger):

        try:

            # get SP of Closed Subtasks of tasks (with subtasks)
            dataset1 = dbOps.getSubtasksStoryPointsClosedInGivenCW_JLR(cursor, initStruct['week_start_date'].strftime('%y-%m-%d %H:%M'), initStruct['week_end_date'].strftime('%y-%m-%d %H:%M'), pd)
            if dataset1[1].iloc[0] != '':
                totalSPofClosedSubtasksJLR = dataset1[1].iloc[0]
            else:
                totalSPofClosedSubtasksJLR = 0

            # get SP of Closed tasks (without subtasks)
            dataset2 = dbOps.getMainTasksStoryPointsClosedInGivenCW_JLR(cursor, initStruct['week_start_date'].strftime('%y-%m-%d %H:%M'), initStruct['week_end_date'].strftime('%y-%m-%d %H:%M'), pd)
            if dataset2[1].iloc[0] != '':
                totalSPofClosedTasksJLR = dataset1[1].iloc[0]
            else:
                totalSPofClosedTasksJLR = 0

            resultDictJLR['ActualTestedReq_SPsum'] = float(totalSPofClosedSubtasksJLR + totalSPofClosedTasksJLR)
            #logger.info("    Actual tested Req. (IST): {}".format(resultDictJLR['ActualTestedReq_SPsum']))

            # get SP of Open Subtasks of tasks (with subtasks)
            dataset3 = dbOps.getSubtasksStoryPointsOpen_JLR(cursor, pd)
            if dataset3[1].iloc[0] != '':
                totalSPofOpenSubtasksJLR = dataset3[1].iloc[0]
            else:
                totalSPofOpenSubtasksJLR = 0

            # get SP of Open tasks (without subtasks)
            dataset4 = dbOps.getMainTasksStoryPointsOpen_JLR(cursor, pd)
            if dataset4[1].iloc[0] != '':
                totalSPofOpenTasksJLR = dataset4[1].iloc[0]
            else:
                totalSPofOpenTasksJLR = 0

            #get SP of Blocked tasks (without subtasks) for JL
            dataset5 = dbOps.getMainTasksStoryPointsBlocked_JLR(cursor, pd)
            if dataset5[1].iloc[0] != '':
                totalSPofBlockedTasksJLR = dataset5[1].iloc[0]
            else:
                totalSPofBlockedTasksJLR = 0

            resultDictJLR['ActualInProgress_SPsum']  = float(totalSPofOpenSubtasksJLR + totalSPofOpenTasksJLR)
            #logger.info("    Actual In Progress (IST): {}".format(resultDictJLR['ActualInProgress_SPsum']))
            resultDictJLR['ActualBlockedReq_SPsum'] = float(totalSPofBlockedTasksJLR)
            #logger.info("    Actual blocked Req. (IST): {}".format(resultDictJLR['ActualBlockedReq_SPsum']))
            resultDictJLR['ActualSummaryReq_SPsum'] = float(resultDictJLR['ActualTestedReq_SPsum'] + resultDictJLR['ActualInProgress_SPsum'] + resultDictJLR['ActualBlockedReq_SPsum'])
            #logger.info("    Actual Summary (IST): {}".format(resultDictJLR['ActualSummaryReq_SPsum']))

    def main(self):
        initStruct = {}
        issueDataDict = {}
        subtaskDataDict = {}
        resultDictOverall = OrderedDict()
        resultDictIDC5 = OrderedDict()
        resultDictMAP = OrderedDict()
        resultDictJLR = OrderedDict()

        fileOps = FileOperations()
        readInput = ReadInputs()
        web = UrlOperations()
        dbOps = DatabaseOperations()

        iniExists = readInput.ParseIniFile(initStruct, fileOps)
        if iniExists == False:
            print ('CRITICAL ERROR: INI file is not present in the same folder as this script resides. Copy it or create it and try again.')
            sys.exit(1)
        # set other environment variables
        logger = Logger.__call__().get_logger(initStruct)
        initStruct['week_start_date'] = datetime.strptime(initStruct['datetime_week_number'] + '-1', "%Y-W%W-%w")
        initStruct['week_end_date'] = initStruct['week_start_date'] + timedelta(days=6)
        initStruct['mapping_file_path'] = pjoin(initStruct['working_folder'], initStruct['mapping_file'])
        initStruct['jira_export_file_path'] = pjoin(initStruct['data_sources_folder'], initStruct['jira_export_folder'])
        initStruct['sqlite_db_path'] = pjoin(initStruct['data_destinations_folder'], initStruct['sqlite_db_filename'])
        # search for the latest exported file from Jira based on filename pattern
        fileNamePattern = initStruct['jira_export_filename_pattern'] + '*.' + initStruct['jira_export_file_extension']
        lastExportFile = fileOps.getLastExportedFileByNamePattern(initStruct['jira_export_file_path'], fileNamePattern)
        initStruct['last_jira_export_file_path'] = pjoin(initStruct['jira_export_file_path'], lastExportFile)

        # check sources and destination availability
        self.CheckExternalAvailability(fileOps, web, initStruct, logger)

        # open data sources, check and extract data then write data to database
        dbColNames = []
        dbColNamesMap = []
        excelColNames = []
        excelColNamesMap = []
        excelColIdxDict = {}

        # check the mapping between jira exported file and db columns
        with open(initStruct['mapping_file_path']) as json_file:
            data = json.load(json_file)
            issueMap = data['issue'][0]
            for key, value in issueMap.items():
                dbColNamesMap.append(key)
                excelColNamesMap.append(value)

        # open data sources, extract data then write data
        excelOps = ExcelOperations()
        activeWorksheet = excelOps.openGivenWorkbookWithActiveWorksheet(initStruct['last_jira_export_file_path'])

        # check the mapping between Excel's columns and dictionary's keys
        for colInfo in activeWorksheet.iter_cols(max_row=1, max_col=activeWorksheet.max_column):
            for colindex in colInfo:
                excelColIdxDict[str(colindex.value)] = str(colindex.col_idx)

        excelColNamesIntersect = set(excelColNames).intersection(excelColNamesMap)
        excelColNamesDiff = set(excelColNamesIntersect).difference(excelColNamesMap)
        if len(excelColNamesDiff) != 0:
            logger.info('Some columns in Excel file are not matching with those declared in the mapping file. Check the Columns in Jira or the mapping file and try again.')
            sys.exit()

        # open database
        sqliteDb = dbOps.openDatabase(initStruct['sqlite_db_path'])
        cursor = sqliteDb.cursor()
        issueColnames = dbOps.getTableColumns(sqliteDb, 'issue')
        # get Projects
        projectData = sqliteDb.execute('SELECT * FROM project')
        dbProjectData = projectData.fetchall()

        # check the mapping between Issue table columns and mapped ones
        for row in issueColnames:
            dbColNames.append(row[0])

        dbColNamesIntersect = set(dbColNames).intersection(dbColNamesMap)
        dbColNamesDiff = set(dbColNamesIntersect).difference(dbColNamesMap)
        if len(dbColNamesDiff) != 0:
            logger.info('Some columns in database are not matching with those declared in the mapping file. Check the database or the mapping file and try again.')
            sys.exit()

        # if mapping is ok then get data and then write it in database
        allRows = activeWorksheet.max_row
        self.ReadDataThenWrite(sqliteDb, cursor, allRows, issueMap, excelColIdxDict, activeWorksheet, issueDataDict, subtaskDataDict, dbProjectData, logger)

# database is now feeded and now starting the calculations

    #TBD: c.execute("alter table issue add column 'SubTasksSP' REAL") - add sum of subtask SP
        logger.info("ETL task has been done ---------------------------------------------------------")
        logger.info("Start calculations for {0}".format(initStruct['easy_week_number'] + '... '))

    # Slide 5
        logger.info("...for Slide 5...")
        self.GetResultsForSlide5(dbOps, cursor, initStruct, resultDictOverall, resultDictIDC5, resultDictJLR, resultDictMAP, logger)

    # Slide 6
        logger.info("...for Slide 6...")
        self.GetResultsForSlide6(dbOps, cursor, initStruct, resultDictIDC5, resultDictJLR, resultDictMAP, logger)

    # Slide 7
        logger.info("...for Slide 7...")
        self.GetResultsForSlide7(dbOps, cursor, initStruct, resultDictOverall, logger)

    # Slide 8
        logger.info("...for Slide 8...")
        self.GetResultsForSlide8(dbOps, cursor, initStruct, resultDictIDC5, logger)

    # Slide 9
        logger.info("...for Slide 9...")
        self.GetResultsForSlide9(dbOps, cursor, initStruct, resultDictMAP, logger)

    # Slide 10
        logger.info("...for Slide 10.")
        self.GetResultsForSlide10(dbOps, cursor, initStruct, resultDictJLR, logger)

    # write results in database
        # get all info before to be written in database
        for project in dbProjectData:
            if project[1] == 'Overall': 
                resultDictOverall['ProjectID'] = project[0]
            elif project[1] == 'DAI DASy IDC5': 
                resultDictIDC5['ProjectID'] = project[0]
            elif project[1] == 'DAI DASy MAP': 
                resultDictMAP['ProjectID'] = project[0]
            elif project[1] == 'DASy JLR L663': 
                resultDictJLR['ProjectID'] = project[0]

        resultDictOverall['CalculationDateTime'] = datetime.today().strftime('%y-%m-%d %H:%M')
        resultDictOverall['CW'] = initStruct['week_number']
        resultDictOverall['CWstartDate'] = initStruct['week_start_date'].strftime('%y-%m-%d')
        resultDictOverall['CWendDate'] = initStruct['week_end_date'].strftime('%y-%m-%d')

        cursor.execute('SELECT * FROM WeeklyReportResults WHERE CW=?', (resultDictOverall["CW"],))
        existingCW = cursor.fetchall()
        if not existingCW:
            dbOps.insertAllResults(sqliteDb, cursor, resultDictOverall, resultDictIDC5, resultDictMAP, resultDictJLR, logger)
        else:
            resultDictOverall['ResultsID'] = resultDictIDC5['ResultsID'] = resultDictMAP['ResultsID'] = resultDictJLR['ResultsID'] = existingCW[0][0]
            dbOps.updateAllResults(sqliteDb, cursor, resultDictOverall, resultDictIDC5, resultDictMAP, resultDictJLR, logger)

    # Write results in a new Excel file
        logger.info("Write results in a new Excel file...")
        newExcelWb = excelOps.openNewWorkbook()
        sheet = newExcelWb.active
        #sheet.title = initStruct['easy_week_number']

        for project in dbProjectData:
            if project[1] == 'Overall': 
                resultDictOverall['ProjectName'] = project[1]
            elif project[1] == 'DAI DASy IDC5': 
                resultDictIDC5['ProjectName'] = project[1]
            elif project[1] == 'DAI DASy MAP': 
                resultDictMAP['ProjectName'] = project[1]
            elif project[1] == 'DASy JLR L663': 
                resultDictJLR['ProjectName'] = project[1]

        resultsExcelFile = excelOps.writeResults(newExcelWb, sheet, initStruct, resultDictOverall, resultDictIDC5, resultDictMAP, resultDictJLR, logger)
        # show the Excel file
        #subprocess.Popen(["start", "/WAIT", resultsExcelFile], shell=True)
        logger.info("Find the easy-to-read results here: " + resultsExcelFile)

        logger.info("Everything done ----------------------------------------------------------------")


# --- Functions -----------------------------------------------------------------------------------

if __name__ == "__main__":
    Main().main()